import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import matplotlib.pyplot as plt
import os
import chardet
import logging
from datetime import datetime, timedelta
from pathlib import Path
import matplotlib.dates as mdates
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

class TemperatureAnalyzer:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Построение графика температур")
        self.root.geometry("800x600")
        
        # Создаем папку для графиков и Excel файлов
        self.output_dir = Path("temperature_analysis")
        self.output_dir.mkdir(exist_ok=True)
        
        # Настройка логирования
        self.setup_logging()
        
        # Создание интерфейса
        self.create_widgets()
        
    def setup_logging(self):
        # Настройка логирования
        log_file = self.output_dir / "analysis.log"
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_file, encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
        
    def create_widgets(self):
        frame = tk.Frame(self.root)
        frame.pack(padx=20, pady=20, fill=tk.BOTH, expand=True)
        
        # Кнопка выбора файлов
        btn_select_files = tk.Button(
            frame,
            text="Выбрать файлы",
            command=self.build_graphs,
            width=20,
            height=2
        )
        btn_select_files.pack(pady=10)
        
        # Область для логов
        self.log_area = scrolledtext.ScrolledText(
            frame,
            height=20,
            width=80
        )
        self.log_area.pack(pady=10, fill=tk.BOTH, expand=True)
        
        # Добавляем обработчик для вывода логов в текстовое поле
        class TextHandler(logging.Handler):
            def __init__(self, text_widget):
                logging.Handler.__init__(self)
                self.text_widget = text_widget

            def emit(self, record):
                msg = self.format(record)
                def append():
                    self.text_widget.configure(state='normal')
                    self.text_widget.insert(tk.END, msg + '\n')
                    self.text_widget.configure(state='disabled')
                    self.text_widget.see(tk.END)
                self.text_widget.after(0, append)

        # Добавляем обработчик в логгер
        text_handler = TextHandler(self.log_area)
        text_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
        self.logger.addHandler(text_handler)

    def detect_encoding(self, file_path):
        """Определяет кодировку файла."""
        with open(file_path, 'rb') as f:
            raw_data = f.read(10000)  # Читаем первые 10Кб для анализа
        result = chardet.detect(raw_data)
        return result['encoding']

    def create_excel_report(self, file_path, times, sp_values, pv1_values, pv2_values, graph_path):
        try:
            # Создаем DataFrame с данными
            df = pd.DataFrame({
                'Время': times,
                'Установленная температура (SP)': sp_values,
                'Зафиксированная температура 1 (PV1)': pv1_values,
                'Зафиксированная температура 2 (PV2)': pv2_values
            })

            # Создаем Excel файл
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            excel_path = self.output_dir / f"{base_name}_report.xlsx"
            
            # Сохраняем DataFrame в Excel
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Данные', index=False)
                
                # Получаем рабочий лист
                worksheet = writer.sheets['Данные']
                
                # Настраиваем стили заголовков
                header_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
                header_font = Font(bold=True)
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                
                # Автоматическая настройка ширины столбцов
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Добавляем график
                img = Image(graph_path)
                worksheet.add_image(img, 'F2')
                
                # Добавляем статистику
                stats_row = len(df) + 3
                worksheet.cell(row=stats_row, column=1, value='Статистика:')
                worksheet.cell(row=stats_row, column=1).font = Font(bold=True)
                
                # Средние значения
                worksheet.cell(row=stats_row + 1, column=1, value='Среднее значение:')
                worksheet.cell(row=stats_row + 1, column=2, value=df['Установленная температура (SP)'].mean())
                worksheet.cell(row=stats_row + 1, column=3, value=df['Зафиксированная температура 1 (PV1)'].mean())
                worksheet.cell(row=stats_row + 1, column=4, value=df['Зафиксированная температура 2 (PV2)'].mean())
                
                # Максимальные значения
                worksheet.cell(row=stats_row + 2, column=1, value='Максимальное значение:')
                worksheet.cell(row=stats_row + 2, column=2, value=df['Установленная температура (SP)'].max())
                worksheet.cell(row=stats_row + 2, column=3, value=df['Зафиксированная температура 1 (PV1)'].max())
                worksheet.cell(row=stats_row + 2, column=4, value=df['Зафиксированная температура 2 (PV2)'].max())
                
                # Минимальные значения
                worksheet.cell(row=stats_row + 3, column=1, value='Минимальное значение:')
                worksheet.cell(row=stats_row + 3, column=2, value=df['Установленная температура (SP)'].min())
                worksheet.cell(row=stats_row + 3, column=3, value=df['Зафиксированная температура 1 (PV1)'].min())
                worksheet.cell(row=stats_row + 3, column=4, value=df['Зафиксированная температура 2 (PV2)'].min())

            self.logger.info(f"Excel отчет сохранен: {excel_path}")
            return excel_path

        except Exception as e:
            self.logger.error(f"Ошибка при создании Excel отчета: {str(e)}")
            return None

    def process_file(self, file_path):
        try:
            self.logger.info(f"Обработка файла: {file_path}")
            
            encoding = self.detect_encoding(file_path)
            if encoding is None:
                raise ValueError("Не удалось определить кодировку файла.")

            times = []
            sp_values = []
            pv1_values = []
            pv2_values = []

            with open(file_path, 'r', encoding=encoding) as file:
                for line in file:
                    parts = line.strip().split()
                    if len(parts) >= 5:
                        time_str = parts[0]
                        try:
                            sp = float(parts[2].strip())
                            pv1 = float(parts[3].strip())
                            pv2 = float(parts[4].strip())
                        except ValueError:
                            continue  # Пропускаем повреждённые строки

                        times.append(time_str)
                        sp_values.append(sp)
                        pv1_values.append(pv1)
                        pv2_values.append(pv2)

            if not times:
                raise ValueError("Файл прочитан, но данные не найдены.")

            # Построение графика
            plt.style.use('bmh')  # Используем встроенный стиль matplotlib
            fig, ax = plt.subplots(figsize=(24, 12))  # Увеличиваем размер графика
            
            # Настройка сетки
            ax.grid(True, linestyle='--', alpha=0.7)
            ax.set_axisbelow(True)  # Сетка под графиками
            
            # Построение линий с улучшенными параметрами
            line1, = ax.plot(times, sp_values, label='Установленная температура (SP)', 
                           color='#1f77b4', linestyle='--', linewidth=3)
            line2, = ax.plot(times, pv1_values, label='Зафиксированная температура 1 (PV1)', 
                           color='#2ca02c', linewidth=3)
            line3, = ax.plot(times, pv2_values, label='Зафиксированная температура 2 (PV2)', 
                           color='#ff7f0e', linewidth=3)

            # Настройка осей
            ax.set_title('График изменения температур во времени', fontsize=18, pad=20)
            ax.set_xlabel('Время', fontsize=14)
            ax.set_ylabel('Температура (°C)', fontsize=14)
            
            # Настройка оси Y (ординат)
            ax.set_ylim(0, 1000)  # Устанавливаем диапазон от 0 до 1000
            ax.yaxis.set_major_locator(plt.MultipleLocator(50))  # Основные деления каждые 50
            ax.yaxis.set_minor_locator(plt.MultipleLocator(12.5))  # Промежуточные деления (3 полоски между основными)
            
            # Настройка оси X (абсцисс)
            # Преобразуем строки времени в объекты datetime
            time_objects = [datetime.strptime(t, '%H:%M:%S') for t in times]
            
            # Создаем список индексов для меток времени (каждые 10 минут)
            tick_indices = []
            tick_labels = []
            current_time = time_objects[0]
            end_time = time_objects[-1]
            
            while current_time <= end_time:
                # Находим ближайший индекс к текущему времени
                idx = min(range(len(time_objects)), 
                         key=lambda i: abs((time_objects[i] - current_time).total_seconds()))
                tick_indices.append(idx)
                tick_labels.append(current_time.strftime('%H:%M'))
                
                # Увеличиваем время на 10 минут используя timedelta
                current_time = current_time + timedelta(minutes=10)
            
            # Устанавливаем метки времени
            ax.set_xticks(tick_indices)
            ax.set_xticklabels(tick_labels, rotation=45, ha='right', fontsize=8)  # Уменьшаем размер шрифта
            
            # Настройка меток времени
            plt.xticks(rotation=45, ha='right', fontsize=8)  # Уменьшаем размер шрифта
            
            # Добавляем сетку для лучшей читаемости
            ax.grid(True, which='major', linestyle='-', alpha=0.7)
            ax.grid(True, which='minor', linestyle=':', alpha=0.4)
            
            # Настройка отступов
            plt.tight_layout()

            # Сохранение графика
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            graph_path = self.output_dir / f"{base_name}_graph.png"
            plt.savefig(graph_path, dpi=600, bbox_inches='tight')  # Увеличиваем разрешение до 600 DPI
            plt.close()

            # Создание Excel отчета
            excel_path = self.create_excel_report(file_path, times, sp_values, pv1_values, pv2_values, graph_path)
            
            if excel_path:
                self.logger.info(f"Обработка файла завершена успешно")
                return True
            else:
                return False

        except Exception as e:
            self.logger.error(f"Ошибка при обработке файла {file_path}: {str(e)}")
            return False

    def build_graphs(self):
        file_paths = filedialog.askopenfilenames(
            title="Выберите файлы с данными",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
        )
        
        if file_paths:
            self.logger.info(f"Выбрано файлов: {len(file_paths)}")
            success_count = 0
            
            for file_path in file_paths:
                if self.process_file(file_path):
                    success_count += 1
            
            self.logger.info(f"Обработка завершена. Успешно обработано файлов: {success_count} из {len(file_paths)}")
            
            if success_count > 0:
                messagebox.showinfo(
                    "Успех",
                    f"Обработано файлов: {success_count} из {len(file_paths)}\n"
                    f"Отчеты сохранены в папку: {self.output_dir}"
                )

    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    app = TemperatureAnalyzer()
    app.run()